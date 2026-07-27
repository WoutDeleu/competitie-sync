#!/usr/bin/env python3
"""competitie-excel — build a "wie wanneer competitie" availability sheet
(.xlsx) for one team's matches on toernooi.nl.

Same interactive flow as competitie_sync.py (sport -> league -> club -> team),
but instead of an .ics it writes an Excel planning grid: one column per match
(Dutch weekday + date + time + "vs opponent"), a Locatie (thuis/uit) row, an
Eten row, and basis/reserve player sections with an availability legend to fill
in by hand.

Requires openpyxl (pip install -r requirements.txt).
"""

from __future__ import annotations

import sys
from datetime import date

from competitie_sync import Session, select_team_matches, slugify

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print(
        "This script needs openpyxl. Install it with:\n"
        "  pip install -r requirements.txt",
        file=sys.stderr,
    )
    sys.exit(1)

# Dutch weekday names, Monday=0 .. Sunday=6 (matching date.weekday()).
DUTCH_WEEKDAYS = [
    "maandag",
    "dinsdag",
    "woensdag",
    "donderdag",
    "vrijdag",
    "zaterdag",
    "zondag",
]

BASIS_ROWS = 8      # player rows in the "basis" (starters) block
RESERVE_ROWS = 7    # player rows in the "reserve" block

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CENTER_V = Alignment(horizontal="center", vertical="center", text_rotation=90)


def _opponent(match, team_name):
    """Return (opponent_name, location_label) for one match."""
    if match.home == team_name:
        return match.away, "thuis"
    if match.away == team_name:
        return match.home, "uit"
    # Fallback: our name didn't match either side exactly.
    return match.away or match.home, "?"


def _header_text(match, opponent):
    """e.g. 'zondag 07/09/2025 19:00 vs Smash For Fun'."""
    weekday = ""
    date_txt = match.date
    if match.date:
        try:
            d, m, y = (int(x) for x in match.date.split("-"))
            weekday = DUTCH_WEEKDAYS[date(y, m, d).weekday()]
            date_txt = f"{d:02d}/{m:02d}/{y}"
        except ValueError:
            pass
    parts = [p for p in (weekday, date_txt, match.time) if p]
    prefix = " ".join(parts)
    vs = f"vs {opponent}" if opponent else "vs ?"
    return f"{prefix} {vs}".strip()


def build_workbook(matches, team_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Blad1"

    red = Font(size=10, color="FFFF0000")
    bold = Font(size=10, bold=True)

    first_col = 3  # column C
    n = len(matches)
    last_col = first_col + n - 1

    # Row 1: match headers (red, wrapped, tall).
    ws.row_dimensions[1].height = 70
    for idx, m in enumerate(matches):
        opp, loc = _opponent(m, team_name)
        col = first_col + idx
        cell = ws.cell(row=1, column=col, value=_header_text(m, opp))
        cell.font = red
        cell.alignment = _CENTER
        cell.border = _BORDER

    # Row 2: Locatie (thuis/uit).
    ws.cell(row=2, column=2, value="Locatie").font = bold
    ws.cell(row=2, column=2).alignment = _CENTER
    for idx, m in enumerate(matches):
        _, loc = _opponent(m, team_name)
        c = ws.cell(row=2, column=first_col + idx, value=loc)
        c.alignment = _CENTER

    # Row 3: Eten (food arrangement, left blank).
    ws.cell(row=3, column=2, value="Eten").font = bold
    ws.cell(row=3, column=2).alignment = _CENTER

    # Player blocks: basis then reserve.
    basis_start = 4
    basis_end = basis_start + BASIS_ROWS - 1
    reserve_start = basis_end + 2  # one blank spacer row between blocks
    reserve_end = reserve_start + RESERVE_ROWS - 1

    _label_block(ws, "basis", basis_start, basis_end)
    _label_block(ws, "reserve", reserve_start, reserve_end)

    # Apply thin borders across the grid (name column + match columns) for the
    # header/location/eten rows and both player blocks.
    for row in [1, 2, 3]:
        _border_row(ws, row, 2, last_col)
    for row in range(basis_start, basis_end + 1):
        _border_row(ws, row, 2, last_col)
    for row in range(reserve_start, reserve_end + 1):
        _border_row(ws, row, 2, last_col)

    # Legend (to the right of the grid).
    legend_col = last_col + 2
    legend = [("x", "kan meedoen"), ("?", "nog niet zeker"), ("", "kan niet meedoen")]
    for i, (mark, meaning) in enumerate(legend):
        ws.cell(row=basis_start + 1 + i, column=legend_col, value=mark)
        ws.cell(row=basis_start + 1 + i, column=legend_col + 1, value=meaning)

    # Column widths and row heights.
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 20.5
    for idx in range(n):
        ws.column_dimensions[get_column_letter(first_col + idx)].width = 11.5
    for r in range(2, reserve_end + 1):
        ws.row_dimensions[r].height = 20

    return wb


def _label_block(ws, label, start, end):
    """Merge column A over the block rows and write a rotated label."""
    ws.merge_cells(
        start_row=start, start_column=1, end_row=end, end_column=1
    )
    cell = ws.cell(row=start, column=1, value=label)
    cell.alignment = _CENTER_V
    cell.font = Font(size=10, bold=True)
    for r in range(start, end + 1):
        ws.cell(row=r, column=1).border = _BORDER


def _border_row(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).border = _BORDER


def main():
    print("competitie-excel — toernooi.nl team availability sheet\n")
    session = Session()

    team_name, team_matches, league = select_team_matches(session)

    wb = build_workbook(team_matches, team_name)
    filename = f"{slugify(team_name)}-wiewanneercomp.xlsx"
    wb.save(filename)

    print(f"\nWrote {filename}: {len(team_matches)} match column(s).")
    print("Open it and fill in the player names and availability marks.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nCancelled.")
        sys.exit(130)
